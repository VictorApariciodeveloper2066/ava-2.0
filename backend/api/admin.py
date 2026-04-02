"""
API Endpoints del Administrador - AVA 2.0
Endpoints para gestión completa del sistema
"""
from flask import Blueprint, request, jsonify, make_response
from flask_jwt_extended import jwt_required, get_jwt_identity
from backend.models import (
    User, Carrera, Asignatura, PeriodoAcademico, Seccion, Inscripcion,
    CodigoProfesor, CodigoAsistencia, AsistenciaManual
)
from backend.extensions import db
from datetime import datetime, date
import random
import string

admin_api_bp = Blueprint('admin_api', __name__)


def check_admin(user_id):
    """Verify user is admin"""
    user = User.query.get(user_id)
    if not user or user.role != 'admin':
        return None
    return user


# =============================================================================
# DASHBOARD
# =============================================================================

@admin_api_bp.route('/dashboard', methods=['GET'])
@jwt_required()
def get_dashboard():
    """
    GET /api/admin/dashboard
    Returns system statistics
    """
    user_id = int(get_jwt_identity())
    admin = check_admin(user_id)
    if not admin:
        return jsonify({"error": "Acceso no autorizado"}), 403
    
    stats = {
        "total_usuarios": User.query.count(),
        "total_estudiantes": User.query.filter_by(role='estudiante').count(),
        "total_profesores": User.query.filter_by(role='profesor').count(),
        "total_admins": User.query.filter_by(role='admin').count(),
        "total_carreras": Carrera.query.filter_by(activo=True).count(),
        "total_materias": Asignatura.query.filter_by(activo=True).count(),
        "total_secciones": Seccion.query.filter_by(activo=True).count(),
        "total_periodos": PeriodoAcademico.query.filter_by(activo=True).count(),
        "codigos_disponibles": CodigoProfesor.query.filter_by(usado=False).count(),
        "codigos_usados": CodigoProfesor.query.filter_by(usado=True).count(),
    }
    
    return jsonify({"stats": stats}), 200


# =============================================================================
# USUARIOS
# =============================================================================

@admin_api_bp.route('/usuarios', methods=['GET'])
@jwt_required()
def get_usuarios():
    """
    GET /api/admin/usuarios
    Returns list of all users with filters
    """
    user_id = int(get_jwt_identity())
    admin = check_admin(user_id)
    if not admin:
        return jsonify({"error": "Acceso no autorizado"}), 403
    
    # Filters
    role = request.args.get('role')
    activo = request.args.get('activo')
    search = request.args.get('search')
    
    query = User.query
    
    if role:
        query = query.filter_by(role=role)
    if activo is not None:
        query = query.filter_by(activo=(activo == 'true'))
    if search:
        query = query.filter(
            db.or_(
                User.username.ilike(f'%{search}%'),
                User.email.ilike(f'%{search}%'),
                User.primer_nombre.ilike(f'%{search}%'),
                User.primer_apellido.ilike(f'%{search}%'),
                User.ci.ilike(f'%{search}%')
            )
        )
    
    usuarios = query.order_by(User.created_at.desc()).all()
    
    return jsonify({
        "usuarios": [
            {
                "id": u.id,
                "username": u.username,
                "email": u.email,
                "nombre": f"{u.primer_nombre or ''} {u.primer_apellido or ''}".strip(),
                "ci": u.ci,
                "role": u.role,
                "activo": u.activo,
                "created_at": u.created_at.isoformat() if u.created_at else None
            }
            for u in usuarios
        ]
    }), 200


@admin_api_bp.route('/usuarios/<int:usuario_id>', methods=['PUT'])
@jwt_required()
def update_usuario(usuario_id):
    """
    PUT /api/admin/usuarios/<id>
    Update user data
    """
    user_id = int(get_jwt_identity())
    admin = check_admin(user_id)
    if not admin:
        return jsonify({"error": "Acceso no autorizado"}), 403
    
    usuario = User.query.get(usuario_id)
    if not usuario:
        return jsonify({"error": "Usuario no encontrado"}), 404
    
    data = request.get_json()
    
    if 'email' in data:
        usuario.email = data['email']
    if 'primer_nombre' in data:
        usuario.primer_nombre = data['primer_nombre']
    if 'primer_apellido' in data:
        usuario.primer_apellido = data['primer_apellido']
    if 'ci' in data:
        usuario.ci = data['ci']
    if 'role' in data:
        usuario.role = data['role']
    if 'activo' in data:
        usuario.activo = data['activo']
    
    db.session.commit()
    
    return jsonify({"message": "Usuario actualizado"}), 200


@admin_api_bp.route('/usuarios/<int:usuario_id>/desactivar', methods=['POST'])
@jwt_required()
def desactivar_usuario(usuario_id):
    """
    POST /api/admin/usuarios/<id>/desactivar
    Deactivate user
    """
    user_id = int(get_jwt_identity())
    admin = check_admin(user_id)
    if not admin:
        return jsonify({"error": "Acceso no autorizado"}), 403
    
    usuario = User.query.get(usuario_id)
    if not usuario:
        return jsonify({"error": "Usuario no encontrado"}), 404
    
    usuario.activo = False
    db.session.commit()
    
    return jsonify({"message": "Usuario desactivado"}), 200


# =============================================================================
# CARRERAS
# =============================================================================

@admin_api_bp.route('/carreras', methods=['GET'])
@jwt_required()
def get_carreras():
    """
    GET /api/admin/carreras
    Returns list of all careers
    """
    user_id = int(get_jwt_identity())
    admin = check_admin(user_id)
    if not admin:
        return jsonify({"error": "Acceso no autorizado"}), 403
    
    carreras = Carrera.query.all()
    
    result = []
    for carrera in carreras:
        # Count subjects
        materias_count = Asignatura.query.filter_by(carrera_id=carrera.id, activo=True).count()
        # Count students
        alumnos_count = User.query.filter_by(carrera_id=carrera.id, role='estudiante', activo=True).count()
        
        result.append({
            "id": carrera.id,
            "nombre": carrera.nombre,
            "codigo": carrera.codigo,
            "duracion_semestres": carrera.duracion_semestres,
            "activo": carrera.activo,
            "materias_count": materias_count,
            "alumnos_count": alumnos_count
        })
    
    return jsonify({"carreras": result}), 200


@admin_api_bp.route('/carreras', methods=['POST'])
@jwt_required()
def create_carrera():
    """
    POST /api/admin/carreras
    Create new career
    """
    user_id = int(get_jwt_identity())
    admin = check_admin(user_id)
    if not admin:
        return jsonify({"error": "Acceso no autorizado"}), 403
    
    data = request.get_json()
    
    carrera = Carrera(
        institucion_id=1,  # Default UNIFA
        nombre=data.get('nombre'),
        codigo=data.get('codigo'),
        duracion_semestres=data.get('duracion_semestres', 10),
        activo=True
    )
    db.session.add(carrera)
    db.session.commit()
    
    return jsonify({"message": "Carrera creada", "id": carrera.id}), 201


@admin_api_bp.route('/carreras/<int:carrera_id>', methods=['PUT'])
@jwt_required()
def update_carrera(carrera_id):
    """
    PUT /api/admin/carreras/<id>
    Update career
    """
    user_id = int(get_jwt_identity())
    admin = check_admin(user_id)
    if not admin:
        return jsonify({"error": "Acceso no autorizado"}), 403
    
    carrera = Carrera.query.get(carrera_id)
    if not carrera:
        return jsonify({"error": "Carrera no encontrada"}), 404
    
    data = request.get_json()
    
    if 'nombre' in data:
        carrera.nombre = data['nombre']
    if 'codigo' in data:
        carrera.codigo = data['codigo']
    if 'duracion_semestres' in data:
        carrera.duracion_semestres = data['duracion_semestres']
    if 'activo' in data:
        carrera.activo = data['activo']
    
    db.session.commit()
    
    return jsonify({"message": "Carrera actualizada"}), 200


# =============================================================================
# ASIGNATURAS (MATERIAS)
# =============================================================================

@admin_api_bp.route('/materias', methods=['GET'])
@jwt_required()
def get_materias():
    """
    GET /api/admin/materias
    Returns list of all subjects
    """
    user_id = int(get_jwt_identity())
    admin = check_admin(user_id)
    if not admin:
        return jsonify({"error": "Acceso no autorizado"}), 403
    
    carrera_id = request.args.get('carrera_id', type=int)
    semestre = request.args.get('semestre', type=int)
    
    query = Asignatura.query
    if carrera_id:
        query = query.filter_by(carrera_id=carrera_id)
    if semestre:
        query = query.filter_by(semestre=semestre)
    
    materias = query.order_by(Asignatura.semestre, Asignatura.nombre).all()
    
    result = []
    for materia in materias:
        carrera = Carrera.query.get(materia.carrera_id)
        secciones_count = Seccion.query.filter_by(asignatura_id=materia.id, activo=True).count()
        
        result.append({
            "id": materia.id,
            "codigo": materia.codigo,
            "nombre": materia.nombre,
            "semestre": materia.semestre,
            "uv": materia.uv,
            "carrera_id": materia.carrera_id,
            "carrera_nombre": carrera.nombre if carrera else "N/A",
            "secciones_count": secciones_count,
            "activo": materia.activo
        })
    
    return jsonify({"materias": result}), 200


@admin_api_bp.route('/materias', methods=['POST'])
@jwt_required()
def create_materia():
    """
    POST /api/admin/materias
    Create new subject
    """
    user_id = int(get_jwt_identity())
    admin = check_admin(user_id)
    if not admin:
        return jsonify({"error": "Acceso no autorizado"}), 403
    
    data = request.get_json()
    
    materia = Asignatura(
        carrera_id=data.get('carrera_id'),
        codigo=data.get('codigo'),
        nombre=data.get('nombre'),
        semestre=data.get('semestre'),
        uv=data.get('uv', 4),
        activo=True
    )
    db.session.add(materia)
    db.session.commit()
    
    return jsonify({"message": "Materia creada", "id": materia.id}), 201


@admin_api_bp.route('/materias/<int:materia_id>', methods=['PUT'])
@jwt_required()
def update_materia(materia_id):
    """
    PUT /api/admin/materias/<id>
    Update subject
    """
    user_id = int(get_jwt_identity())
    admin = check_admin(user_id)
    if not admin:
        return jsonify({"error": "Acceso no autorizado"}), 403
    
    materia = Asignatura.query.get(materia_id)
    if not materia:
        return jsonify({"error": "Materia no encontrada"}), 404
    
    data = request.get_json()
    
    if 'nombre' in data:
        materia.nombre = data['nombre']
    if 'codigo' in data:
        materia.codigo = data['codigo']
    if 'semestre' in data:
        materia.semestre = data['semestre']
    if 'uv' in data:
        materia.uv = data['uv']
    if 'activo' in data:
        materia.activo = data['activo']
    
    db.session.commit()
    
    return jsonify({"message": "Materia actualizada"}), 200


# =============================================================================
# SECCIONES
# =============================================================================

@admin_api_bp.route('/secciones', methods=['GET'])
@jwt_required()
def get_secciones():
    """
    GET /api/admin/secciones
    Returns list of all sections
    """
    user_id = int(get_jwt_identity())
    admin = check_admin(user_id)
    if not admin:
        return jsonify({"error": "Acceso no autorizado"}), 403
    
    materia_id = request.args.get('materia_id', type=int)
    
    query = Seccion.query
    if materia_id:
        query = query.filter_by(asignatura_id=materia_id)
    
    secciones = query.order_by(Seccion.asignatura_id, Seccion.dia).all()
    
    result = []
    for seccion in secciones:
        materia = Asignatura.query.get(seccion.asignatura_id)
        profesor = User.query.get(seccion.profesor_id) if seccion.profesor_id else None
        alumnos_count = Inscripcion.query.filter_by(seccion_id=seccion.id, estado='activa').count()
        
        day_names = ['', 'Lun', 'Mar', 'Mié', 'Jue', 'Vie', 'Sáb', 'Dom']
        
        result.append({
            "id": seccion.id,
            "codigo": seccion.codigo,
            "asignatura_id": seccion.asignatura_id,
            "asignatura_nombre": materia.nombre if materia else "N/A",
            "asignatura_codigo": materia.codigo if materia else "N/A",
            "profesor_id": seccion.profesor_id,
            "profesor_nombre": f"{profesor.primer_nombre or ''} {profesor.primer_apellido or ''}".strip() if profesor else "Sin asignar",
            "dia": seccion.dia,
            "dia_nombre": day_names[seccion.dia] if 0 <= seccion.dia <= 7 else "N/A",
            "start_time": seccion.start_time.strftime("%H:%M") if seccion.start_time else None,
            "end_time": seccion.end_time.strftime("%H:%M") if seccion.end_time else None,
            "aula": seccion.aula,
            "capacidad": seccion.capacidad,
            "alumnos_count": alumnos_count,
            "activo": seccion.activo
        })
    
    return jsonify({"secciones": result}), 200


@admin_api_bp.route('/secciones/filtradas', methods=['GET'])
@jwt_required()
def get_secciones_filtradas():
    """
    GET /api/admin/secciones/filtradas?carrera_id=X&profesor_id=Y&search=Z
    Returns filtered sections list with advanced filters
    """
    user_id = int(get_jwt_identity())
    admin = check_admin(user_id)
    if not admin:
        return jsonify({"error": "Acceso no autorizado"}), 403
    
    carrera_id = request.args.get('carrera_id', type=int)
    profesor_id = request.args.get('profesor_id', type=int)
    search = request.args.get('search')
    
    # Base query
    query = Seccion.query.filter_by(activo=True)
    
    # Filter by carrera (through asignatura)
    if carrera_id:
        query = query.join(Asignatura).filter(Asignatura.carrera_id == carrera_id)
    
    # Filter by profesor
    if profesor_id:
        query = query.filter_by(profesor_id=profesor_id)
    
    # Search filter
    if search:
        query = query.join(Asignatura).filter(
            db.or_(
                Seccion.codigo.ilike(f'%{search}%'),
                Asignatura.nombre.ilike(f'%{search}%'),
                Asignatura.codigo.ilike(f'%{search}%')
            )
        )
    
    secciones = query.order_by(Asignatura.nombre, Seccion.dia).all()
    
    day_names = ['', 'Lun', 'Mar', 'Mié', 'Jue', 'Vie', 'Sáb', 'Dom']
    
    result = []
    for seccion in secciones:
        materia = Asignatura.query.get(seccion.asignatura_id)
        profesor = User.query.get(seccion.profesor_id) if seccion.profesor_id else None
        carrera = Carrera.query.get(materia.carrera_id) if materia else None
        alumnos_count = Inscripcion.query.filter_by(seccion_id=seccion.id, estado='activa').count()
        
        result.append({
            "id": seccion.id,
            "codigo": seccion.codigo,
            "asignatura_id": seccion.asignatura_id,
            "asignatura_codigo": materia.codigo if materia else "N/A",
            "asignatura_nombre": materia.nombre if materia else "N/A",
            "carrera_nombre": carrera.nombre if carrera else "N/A",
            "carrera_id": carrera.id if carrera else None,
            "profesor_id": seccion.profesor_id,
            "profesor_nombre": f"{profesor.primer_nombre or ''} {profesor.primer_apellido or ''}".strip() if profesor else "Sin asignar",
            "dia": seccion.dia,
            "dia_nombre": day_names[seccion.dia] if 0 <= seccion.dia <= 7 else "N/A",
            "start_time": seccion.start_time.strftime("%H:%M") if seccion.start_time else None,
            "end_time": seccion.end_time.strftime("%H:%M") if seccion.end_time else None,
            "aula": seccion.aula,
            "capacidad": seccion.capacidad,
            "alumnos_count": alumnos_count
        })
    
    return jsonify({"secciones": result}), 200


@admin_api_bp.route('/carreras/lista', methods=['GET'])
@jwt_required()
def get_carreras_lista():
    """
    GET /api/admin/carreras/lista
    Returns simplified list of careers for filters
    """
    user_id = int(get_jwt_identity())
    admin = check_admin(user_id)
    if not admin:
        return jsonify({"error": "Acceso no autorizado"}), 403
    
    carreras = Carrera.query.filter_by(activo=True).order_by(Carrera.nombre).all()
    
    return jsonify({
        "carreras": [
            {"id": c.id, "nombre": c.nombre, "codigo": c.codigo}
            for c in carreras
        ]
    }), 200


@admin_api_bp.route('/profesores/lista', methods=['GET'])
@jwt_required()
def get_profesores_lista():
    """
    GET /api/admin/profesores/lista
    Returns simplified list of professors for filters
    """
    user_id = int(get_jwt_identity())
    admin = check_admin(user_id)
    if not admin:
        return jsonify({"error": "Acceso no autorizado"}), 403
    
    profesores = User.query.filter_by(role='profesor', activo=True).order_by(User.primer_nombre).all()
    
    return jsonify({
        "profesores": [
            {
                "id": p.id,
                "nombre": f"{p.primer_nombre or ''} {p.primer_apellido or ''}".strip() or p.username
            }
            for p in profesores
        ]
    }), 200


@admin_api_bp.route('/secciones', methods=['POST'])
@jwt_required()
def create_seccion():
    """
    POST /api/admin/secciones
    Create new section
    """
    user_id = int(get_jwt_identity())
    admin = check_admin(user_id)
    if not admin:
        return jsonify({"error": "Acceso no autorizado"}), 403
    
    data = request.get_json()
    
    # Get period
    periodo = PeriodoAcademico.query.filter_by(es_actual=True, activo=True).first()
    if not periodo:
        return jsonify({"error": "No hay período activo"}), 400
    
    # Parse times
    start_time = datetime.strptime(data.get('start_time', '08:00'), '%H:%M').time()
    end_time = datetime.strptime(data.get('end_time', '10:00'), '%H:%M').time()
    
    seccion = Seccion(
        codigo=data.get('codigo'),
        asignatura_id=data.get('asignatura_id'),
        periodo_id=periodo.id,
        profesor_id=data.get('profesor_id'),
        aula=data.get('aula', 'Por asignar'),
        dia=data.get('dia'),
        start_time=start_time,
        end_time=end_time,
        capacidad=data.get('capacidad', 30),
        activo=True
    )
    db.session.add(seccion)
    db.session.commit()
    
    return jsonify({"message": "Sección creada", "id": seccion.id}), 201


@admin_api_bp.route('/secciones/<int:seccion_id>', methods=['PUT'])
@jwt_required()
def update_seccion(seccion_id):
    """
    PUT /api/admin/secciones/<id>
    Update section (schedule, professor, etc)
    """
    user_id = int(get_jwt_identity())
    admin = check_admin(user_id)
    if not admin:
        return jsonify({"error": "Acceso no autorizado"}), 403
    
    seccion = Seccion.query.get(seccion_id)
    if not seccion:
        return jsonify({"error": "Sección no encontrada"}), 404
    
    data = request.get_json()
    
    if 'dia' in data:
        seccion.dia = data['dia']
    if 'start_time' in data:
        seccion.start_time = datetime.strptime(data['start_time'], '%H:%M').time()
    if 'end_time' in data:
        seccion.end_time = datetime.strptime(data['end_time'], '%H:%M').time()
    if 'aula' in data:
        seccion.aula = data['aula']
    if 'profesor_id' in data:
        seccion.profesor_id = data['profesor_id']
    if 'capacidad' in data:
        seccion.capacidad = data['capacidad']
    if 'activo' in data:
        seccion.activo = data['activo']
    
    db.session.commit()
    
    return jsonify({"message": "Sección actualizada"}), 200


# =============================================================================
# PERÍODOS
# =============================================================================

@admin_api_bp.route('/periodos', methods=['GET'])
@jwt_required()
def get_periodos():
    """
    GET /api/admin/periodos
    Returns list of all academic periods
    """
    user_id = int(get_jwt_identity())
    admin = check_admin(user_id)
    if not admin:
        return jsonify({"error": "Acceso no autorizado"}), 403
    
    periodos = PeriodoAcademico.query.order_by(PeriodoAcademico.fecha_inicio.desc()).all()
    
    return jsonify({
        "periodos": [
            {
                "id": p.id,
                "nombre": p.nombre,
                "fecha_inicio": p.fecha_inicio.isoformat() if p.fecha_inicio else None,
                "fecha_fin": p.fecha_fin.isoformat() if p.fecha_fin else None,
                "activo": p.activo,
                "es_actual": p.es_actual
            }
            for p in periodos
        ]
    }), 200


@admin_api_bp.route('/periodos', methods=['POST'])
@jwt_required()
def create_periodo():
    """
    POST /api/admin/periodos
    Create new academic period
    """
    user_id = int(get_jwt_identity())
    admin = check_admin(user_id)
    if not admin:
        return jsonify({"error": "Acceso no autorizado"}), 403
    
    data = request.get_json()
    
    # If setting as current, deactivate other current periods
    if data.get('es_actual'):
        PeriodoAcademico.query.filter_by(es_actual=True).update({'es_actual': False})
    
    periodo = PeriodoAcademico(
        institucion_id=1,
        nombre=data.get('nombre'),
        fecha_inicio=datetime.strptime(data.get('fecha_inicio'), '%Y-%m-%d').date(),
        fecha_fin=datetime.strptime(data.get('fecha_fin'), '%Y-%m-%d').date(),
        activo=True,
        es_actual=data.get('es_actual', False)
    )
    db.session.add(periodo)
    db.session.commit()
    
    return jsonify({"message": "Período creado", "id": periodo.id}), 201


# =============================================================================
# CÓDIGOS DE PROFESOR
# =============================================================================

@admin_api_bp.route('/codigos-profesor', methods=['GET'])
@jwt_required()
def get_codigos_profesor():
    """
    GET /api/admin/codigos-profesor
    Returns list of professor codes
    """
    user_id = int(get_jwt_identity())
    admin = check_admin(user_id)
    if not admin:
        return jsonify({"error": "Acceso no autorizado"}), 403
    
    codigos = CodigoProfesor.query.order_by(CodigoProfesor.created_at.desc()).all()
    
    result = []
    for codigo in codigos:
        usado_por = User.query.get(codigo.usado_por) if codigo.usado_por else None
        
        result.append({
            "id": codigo.id,
            "codigo": codigo.codigo,
            "usado": codigo.usado,
            "usado_por_nombre": f"{usado_por.primer_nombre or ''} {usado_por.primer_apellido or ''}".strip() if usado_por else None,
            "created_at": codigo.created_at.isoformat() if codigo.created_at else None
        })
    
    return jsonify({
        "codigos": result,
        "disponibles": CodigoProfesor.query.filter_by(usado=False).count(),
        "usados": CodigoProfesor.query.filter_by(usado=True).count()
    }), 200


@admin_api_bp.route('/codigos-profesor', methods=['POST'])
@jwt_required()
def create_codigos_profesor():
    """
    POST /api/admin/codigos-profesor
    Body: { "cantidad": 5 }
    Generate new professor codes
    """
    user_id = int(get_jwt_identity())
    admin = check_admin(user_id)
    if not admin:
        return jsonify({"error": "Acceso no autorizado"}), 403
    
    data = request.get_json()
    cantidad = data.get('cantidad', 1)
    
    created = []
    for _ in range(cantidad):
        # Generate unique 4-character code
        while True:
            code = ''.join(random.choices(string.ascii_uppercase + string.digits, k=4))
            if not CodigoProfesor.query.filter_by(codigo=code).first():
                break
        
        codigo = CodigoProfesor(codigo=code, usado=False)
        db.session.add(codigo)
        created.append(code)
    
    db.session.commit()
    
    return jsonify({
        "message": f"{cantidad} códigos creados",
        "codigos": created
    }), 201


@admin_api_bp.route('/codigos-profesor/<int:codigo_id>', methods=['DELETE'])
@jwt_required()
def delete_codigo_profesor(codigo_id):
    """
    DELETE /api/admin/codigos-profesor/<id>
    Delete unused professor code
    """
    user_id = int(get_jwt_identity())
    admin = check_admin(user_id)
    if not admin:
        return jsonify({"error": "Acceso no autorizado"}), 403
    
    codigo = CodigoProfesor.query.get(codigo_id)
    if not codigo:
        return jsonify({"error": "Código no encontrado"}), 404
    
    if codigo.usado:
        return jsonify({"error": "No se puede eliminar un código usado"}), 400
    
    db.session.delete(codigo)
    db.session.commit()
    
    return jsonify({"message": "Código eliminado"}), 200


# =============================================================================
# ASIGNAR ALUMNOS
# =============================================================================

@admin_api_bp.route('/alumnos/disponibles', methods=['GET'])
@jwt_required()
def get_alumnos_disponibles():
    """
    GET /api/admin/alumnos/disponibles?seccion_id=X
    Returns students not enrolled in section
    """
    user_id = int(get_jwt_identity())
    admin = check_admin(user_id)
    if not admin:
        return jsonify({"error": "Acceso no autorizado"}), 403
    
    seccion_id = request.args.get('seccion_id', type=int)
    
    # Get all students
    alumnos = User.query.filter_by(role='estudiante', activo=True).all()
    
    # Get already enrolled students
    enrolled_ids = []
    if seccion_id:
        inscripciones = Inscripcion.query.filter_by(seccion_id=seccion_id, estado='activa').all()
        enrolled_ids = [i.estudiante_id for i in inscripciones]
    
    result = []
    for alumno in alumnos:
        if alumno.id not in enrolled_ids:
            result.append({
                "id": alumno.id,
                "nombre": f"{alumno.primer_nombre or ''} {alumno.primer_apellido or ''}".strip() or alumno.username,
                "ci": alumno.ci,
                "inscrito": alumno.id in enrolled_ids
            })
    
    return jsonify({"alumnos": result}), 200


@admin_api_bp.route('/alumnos/asignar', methods=['POST'])
@jwt_required()
def asignar_alumnos():
    """
    POST /api/admin/alumnos/asignar
    Body: { "seccion_id": 123, "alumno_ids": [1, 2, 3] }
    Assign students to section
    """
    user_id = int(get_jwt_identity())
    admin = check_admin(user_id)
    if not admin:
        return jsonify({"error": "Acceso no autorizado"}), 403
    
    data = request.get_json()
    seccion_id = data.get('seccion_id')
    alumno_ids = data.get('alumno_ids', [])
    
    if not seccion_id:
        return jsonify({"error": "seccion_id requerido"}), 400
    
    seccion = Seccion.query.get(seccion_id)
    if not seccion:
        return jsonify({"error": "Sección no encontrada"}), 404
    
    # Check capacity
    current_count = Inscripcion.query.filter_by(seccion_id=seccion_id, estado='activa').count()
    available = seccion.capacidad - current_count
    
    if len(alumno_ids) > available:
        return jsonify({"error": f"Solo hay {available} cupos disponibles"}), 400
    
    # Create enrollments
    created = 0
    for alumno_id in alumno_ids:
        # Check if already enrolled
        existing = Inscripcion.query.filter_by(
            estudiante_id=alumno_id,
            seccion_id=seccion_id,
            estado='activa'
        ).first()
        
        if not existing:
            inscripcion = Inscripcion(
                estudiante_id=alumno_id,
                seccion_id=seccion_id,
                estado='activa'
            )
            db.session.add(inscripcion)
            created += 1
    
    db.session.commit()
    
    return jsonify({"message": f"{created} alumnos asignados"}), 201


# =============================================================================
# HISTORIAL
# =============================================================================

@admin_api_bp.route('/historial/<int:alumno_id>', methods=['GET'])
@jwt_required()
def get_historial_completo(alumno_id):
    """
    GET /api/admin/historial/<alumno_id>
    Returns complete attendance history for a student
    """
    user_id = int(get_jwt_identity())
    admin = check_admin(user_id)
    if not admin:
        return jsonify({"error": "Acceso no autorizado"}), 403
    
    alumno = User.query.get(alumno_id)
    if not alumno:
        return jsonify({"error": "Alumno no encontrado"}), 404
    
    # Get all inscriptions
    inscripciones = Inscripcion.query.filter_by(estudiante_id=alumno_id, estado='activa').all()
    
    historial = []
    for insc in inscripciones:
        seccion = Seccion.query.get(insc.seccion_id)
        if not seccion:
            continue
        
        materia = Asignatura.query.get(seccion.asignatura_id)
        
        # Get attendance records
        asistencias = AsistenciaManual.query.filter_by(
            seccion_id=seccion.id,
            alumno_id=alumno_id
        ).order_by(AsistenciaManual.fecha.desc()).all()
        
        presentes = sum(1 for a in asistencias if a.estado == 'presente')
        total = len(asistencias)
        porcentaje = (presentes / total * 100) if total > 0 else 100
        
        historial.append({
            "seccion_id": seccion.id,
            "materia_codigo": materia.codigo if materia else "N/A",
            "materia_nombre": materia.nombre if materia else "N/A",
            "total_clases": total,
            "presentes": presentes,
            "ausentes": sum(1 for a in asistencias if a.estado == 'ausente'),
            "justificados": sum(1 for a in asistencias if a.estado == 'justificado'),
            "porcentaje": round(porcentaje, 1),
            "registros": [
                {
                    "fecha": a.fecha.isoformat(),
                    "estado": a.estado
                }
                for a in asistencias
            ]
        })
    
    return jsonify({
        "alumno": {
            "id": alumno.id,
            "nombre": f"{alumno.primer_nombre or ''} {alumno.primer_apellido or ''}".strip() or alumno.username,
            "ci": alumno.ci
        },
        "historial": historial
    }), 200


# =============================================================================
# REPORTES
# =============================================================================

@admin_api_bp.route('/reporte/<report_type>', methods=['GET'])
@jwt_required()
def download_report(report_type):
    """
    GET /api/admin/reporte/<report_type>?format=pdf|excel
    Download system reports
    """
    user_id = int(get_jwt_identity())
    admin = check_admin(user_id)
    if not admin:
        return jsonify({"error": "Acceso no autorizado"}), 403
    
    format_type = request.args.get('format', 'pdf')
    
    if report_type == 'usuarios':
        return generate_usuarios_report(format_type)
    elif report_type == 'asistencia':
        return generate_asistencia_report(format_type)
    else:
        return jsonify({"error": "Tipo de reporte no válido"}), 400


def generate_usuarios_report(format_type):
    """Generate users report"""
    usuarios = User.query.all()
    
    if format_type == 'excel':
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from io import BytesIO
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Usuarios"
        
        headers = ['ID', 'Username', 'Email', 'Nombre', 'CI', 'Rol', 'Activo']
        ws.append(headers)
        
        header_fill = PatternFill(start_color='39E079', end_color='39E079', fill_type='solid')
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        for u in usuarios:
            ws.append([
                u.id,
                u.username,
                u.email,
                f"{u.primer_nombre or ''} {u.primer_apellido or ''}".strip(),
                u.ci,
                u.role,
                'Sí' if u.activo else 'No'
            ])
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = make_response(buffer.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = 'attachment; filename=usuarios.xlsx'
        return response
    
    else:  # PDF
        from reportlab.lib.pagesizes import letter
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
        from reportlab.lib.styles import getSampleStyleSheet
        from io import BytesIO
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        styles = getSampleStyles()
        
        elements.append(Paragraph("Reporte de Usuarios", styles['Heading1']))
        
        data = [['ID', 'Username', 'Nombre', 'CI', 'Rol', 'Estado']]
        for u in usuarios:
            data.append([
                str(u.id),
                u.username,
                f"{u.primer_nombre or ''} {u.primer_apellido or ''}".strip(),
                u.ci or 'N/A',
                u.role,
                'Activo' if u.activo else 'Inactivo'
            ])
        
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#39E079')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ]))
        elements.append(table)
        
        doc.build(elements)
        buffer.seek(0)
        
        response = make_response(buffer.getvalue())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = 'attachment; filename=usuarios.pdf'
        return response


def generate_asistencia_report(format_type):
    """Generate attendance report"""
    # Implementation similar to users report
    return jsonify({"message": "Reporte de asistencia"}), 200
