"""
API Endpoints del Profesor - AVA 2.0
Endpoints para funcionalidades del profesor: materias, alumnos, asistencia, reportes
"""
from flask import Blueprint, request, jsonify, make_response, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity
from backend.models import (
    User, Seccion, Inscripcion, Asignatura,
    CodigoProfesor, CodigoAsistencia, AsistenciaManual,
    LogAsistenciaProfesor, PeriodoAcademico
)
from backend.extensions import db
from datetime import datetime, date, timedelta
import random
import string

profesor_api_bp = Blueprint('profesor_api', __name__)


@profesor_api_bp.route('/validar-codigo', methods=['POST'])
def validar_codigo_profesor():
    """
    POST /api/profesor/validar-codigo
    Body: { "codigo": "PROF-1234" }
    Valida si un código de profesor es válido y no usado
    """
    data = request.get_json()
    codigo = data.get('codigo', '').upper().strip()
    
    if not codigo:
        return jsonify({"error": "Código requerido"}), 400
    
    # Formato: 4 caracteres alfanuméricos
    if len(codigo) != 4:
        return jsonify({"error": "El código debe tener 4 caracteres"}), 400
    
    # Buscar código
    codigo_obj = CodigoProfesor.query.filter_by(codigo=codigo).first()
    
    if not codigo_obj:
        return jsonify({"error": "Código inválido"}), 404
    
    if codigo_obj.usado:
        return jsonify({"error": "Este código ya fue utilizado"}), 400
    
    return jsonify({
        "valid": True,
        "message": "Código válido"
    }), 200


@profesor_api_bp.route('/materias', methods=['GET'])
@jwt_required()
def get_materias_profesor():
    """
    GET /api/profesor/materias
    Returns list of subjects the professor teaches
    """
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    
    if not user or user.role != 'profesor':
        return jsonify({"error": "Acceso no autorizado"}), 403
    
    # Get sections where user is professor
    secciones = Seccion.query.filter_by(
        profesor_id=user_id,
        activo=True
    ).all()
    
    # Group by asignatura
    materias_dict = {}
    for seccion in secciones:
        asignatura = Asignatura.query.get(seccion.asignatura_id)
        if not asignatura:
            continue
        
        key = asignatura.id
        if key not in materias_dict:
            # Count enrolled students
            alumnos_count = Inscripcion.query.filter_by(
                seccion_id=seccion.id,
                estado='activa'
            ).count()
            
            materias_dict[key] = {
                "asignatura_id": asignatura.id,
                "codigo": asignatura.codigo,
                "nombre": asignatura.nombre,
                "semestre": asignatura.semestre,
                "secciones": [],
                "total_alumnos": 0
            }
        
        # Count students for this section
        alumnos_seccion = Inscripcion.query.filter_by(
            seccion_id=seccion.id,
            estado='activa'
        ).count()
        
        materias_dict[key]["secciones"].append({
            "id": seccion.id,
            "codigo": seccion.codigo,
            "dia": seccion.dia,
            "start_time": seccion.start_time.strftime("%H:%M") if seccion.start_time else None,
            "end_time": seccion.end_time.strftime("%H:%M") if seccion.end_time else None,
            "aula": seccion.aula,
            "alumnos_count": alumnos_seccion
        })
        materias_dict[key]["total_alumnos"] += alumnos_seccion
    
    return jsonify({
        "materias": list(materias_dict.values())
    }), 200


@profesor_api_bp.route('/seccion/<int:seccion_id>/alumnos', methods=['GET'])
@jwt_required()
def get_alumnos_seccion(seccion_id):
    """
    GET /api/profesor/seccion/<seccion_id>/alumnos
    Returns list of students enrolled in a section
    """
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    
    if not user or user.role != 'profesor':
        return jsonify({"error": "Acceso no autorizado"}), 403
    
    seccion = Seccion.query.get(seccion_id)
    if not seccion:
        return jsonify({"error": "Sección no encontrada"}), 404
    
    if seccion.profesor_id != user_id:
        return jsonify({"error": "No eres profesor de esta sección"}), 403
    
    # Get enrolled students
    inscripciones = Inscripcion.query.filter_by(
        seccion_id=seccion_id,
        estado='activa'
    ).all()
    
    alumnos = []
    for insc in inscripciones:
        alumno = User.query.get(insc.estudiante_id)
        if alumno:
            # Get attendance stats
            total_clases = AsistenciaManual.query.filter_by(
                seccion_id=seccion_id,
                alumno_id=alumno.id
            ).count()
            
            presentes = AsistenciaManual.query.filter_by(
                seccion_id=seccion_id,
                alumno_id=alumno.id,
                estado='presente'
            ).count()
            
            porcentaje = (presentes / total_clases * 100) if total_clases > 0 else 100
            
            alumnos.append({
                "id": alumno.id,
                "nombre": f"{alumno.primer_nombre or ''} {alumno.primer_apellido or ''}".strip() or alumno.username,
                "ci": alumno.ci,
                "email": alumno.email,
                "asistencia_porcentaje": round(porcentaje, 1),
                "total_clases": total_clases,
                "presentes": presentes
            })
    
    # Get today's attendance
    today = date.today()
    asistencias_hoy = {}
    for asistencia in AsistenciaManual.query.filter_by(seccion_id=seccion_id, fecha=today).all():
        asistencias_hoy[asistencia.alumno_id] = asistencia.estado
    
    # Add today's status to each alumno
    for alumno in alumnos:
        alumno["estado_hoy"] = asistencias_hoy.get(alumno["id"], "sin_marcar")
    
    return jsonify({
        "seccion": {
            "id": seccion.id,
            "codigo": seccion.codigo,
            "dia": seccion.dia,
            "start_time": seccion.start_time.strftime("%H:%M") if seccion.start_time else None,
            "end_time": seccion.end_time.strftime("%H:%M") if seccion.end_time else None,
            "aula": seccion.aula
        },
        "alumnos": alumnos,
        "fecha": today.isoformat()
    }), 200


@profesor_api_bp.route('/generar-codigo', methods=['POST'])
@jwt_required()
def generar_codigo_asistencia():
    """
    POST /api/profesor/generar-codigo
    Body: { "seccion_id": 123 }
    Generates a 4-digit attendance code that lasts the entire class
    """
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    
    if not user or user.role != 'profesor':
        return jsonify({"error": "Acceso no autorizado"}), 403
    
    data = request.get_json()
    seccion_id = data.get('seccion_id')
    
    if not seccion_id:
        return jsonify({"error": "seccion_id requerido"}), 400
    
    seccion = Seccion.query.get(seccion_id)
    if not seccion:
        return jsonify({"error": "Sección no encontrada"}), 404
    
    if seccion.profesor_id != user_id:
        return jsonify({"error": "No eres profesor de esta sección"}), 403
    
    # Calculate expiration (end of class time)
    today = date.today()
    if seccion.end_time:
        expira = datetime.combine(today, seccion.end_time)
    else:
        expira = datetime.now() + timedelta(hours=2)  # Default 2 hours
    
    # Generate 4-digit code
    codigo = ''.join(random.choices(string.ascii_uppercase + string.digits, k=4))
    
    # Invalidate previous codes for this section today
    CodigoAsistencia.query.filter_by(
        seccion_id=seccion_id,
        fecha=today
    ).delete()
    
    # Create new code
    codigo_obj = CodigoAsistencia(
        seccion_id=seccion_id,
        codigo=codigo,
        fecha=today,
        expira=expira,
        created_by=user_id
    )
    db.session.add(codigo_obj)
    db.session.commit()
    
    return jsonify({
        "codigo": codigo,
        "expira": expira.isoformat(),
        "seccion_id": seccion_id
    }), 201


@profesor_api_bp.route('/marcar-asistencia', methods=['POST'])
@jwt_required()
def marcar_asistencia_manual():
    """
    POST /api/profesor/marcar-asistencia
    Body: { "seccion_id": 123, "alumno_id": 456, "estado": "presente" }
    Marks attendance manually for a student
    """
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    
    if not user or user.role != 'profesor':
        return jsonify({"error": "Acceso no autorizado"}), 403
    
    data = request.get_json()
    seccion_id = data.get('seccion_id')
    alumno_id = data.get('alumno_id')
    estado = data.get('estado', 'presente')
    
    if not seccion_id or not alumno_id:
        return jsonify({"error": "seccion_id y alumno_id requeridos"}), 400
    
    if estado not in ['presente', 'ausente', 'justificado']:
        return jsonify({"error": "Estado inválido"}), 400
    
    seccion = Seccion.query.get(seccion_id)
    if not seccion:
        return jsonify({"error": "Sección no encontrada"}), 404
    
    if seccion.profesor_id != user_id:
        return jsonify({"error": "No eres profesor de esta sección"}), 403
    
    # Check if student is enrolled
    inscripcion = Inscripcion.query.filter_by(
        estudiante_id=alumno_id,
        seccion_id=seccion_id,
        estado='activa'
    ).first()
    
    if not inscripcion:
        return jsonify({"error": "El alumno no está inscrito en esta sección"}), 400
    
    # Get or create attendance record
    today = date.today()
    asistencia = AsistenciaManual.query.filter_by(
        seccion_id=seccion_id,
        alumno_id=alumno_id,
        fecha=today
    ).first()
    
    estado_anterior = None
    if asistencia:
        estado_anterior = asistencia.estado
        asistencia.estado = estado
        asistencia.marcado_por = user_id
    else:
        asistencia = AsistenciaManual(
            seccion_id=seccion_id,
            alumno_id=alumno_id,
            estado=estado,
            fecha=today,
            marcado_por=user_id
        )
        db.session.add(asistencia)
    
    db.session.flush()
    
    # Log the change
    log = LogAsistenciaProfesor(
        asistencia_id=asistencia.id,
        alumno_id=alumno_id,
        seccion_id=seccion_id,
        estado_anterior=estado_anterior,
        estado_nuevo=estado,
        modificado_por=user_id
    )
    db.session.add(log)
    db.session.commit()
    
    return jsonify({
        "message": "Asistencia actualizada",
        "alumno_id": alumno_id,
        "estado": estado
    }), 200


@profesor_api_bp.route('/historial/<int:alumno_id>/<int:seccion_id>', methods=['GET'])
@jwt_required()
def get_historial_alumno(alumno_id, seccion_id):
    """
    GET /api/profesor/historial/<alumno_id>/<seccion_id>
    Returns attendance history for a student in a specific section
    """
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    
    if not user or user.role != 'profesor':
        return jsonify({"error": "Acceso no autorizado"}), 403
    
    seccion = Seccion.query.get(seccion_id)
    if not seccion or seccion.profesor_id != user_id:
        return jsonify({"error": "No tienes acceso a esta sección"}), 403
    
    # Get attendance records
    asistencias = AsistenciaManual.query.filter_by(
        seccion_id=seccion_id,
        alumno_id=alumno_id
    ).order_by(AsistenciaManual.fecha.desc()).all()
    
    # Get logs for changes
    logs = LogAsistenciaProfesor.query.filter_by(
        seccion_id=seccion_id,
        alumno_id=alumno_id
    ).order_by(LogAsistenciaProfesor.fecha.desc()).all()
    
    historial = []
    for asistencia in asistencias:
        historial.append({
            "fecha": asistencia.fecha.isoformat(),
            "estado": asistencia.estado,
            "created_at": asistencia.created_at.isoformat()
        })
    
    cambios = []
    for log in logs:
        cambios.append({
            "fecha": log.fecha.isoformat(),
            "estado_anterior": log.estado_anterior,
            "estado_nuevo": log.estado_nuevo
        })
    
    return jsonify({
        "alumno_id": alumno_id,
        "seccion_id": seccion_id,
        "historial": historial,
        "cambios": cambios
    }), 200


@profesor_api_bp.route('/reporte/<int:seccion_id>/pdf', methods=['GET'])
@jwt_required()
def descargar_reporte_pdf(seccion_id):
    """
    GET /api/profesor/reporte/<seccion_id>/pdf
    Downloads PDF attendance report
    """
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    
    if not user or user.role != 'profesor':
        return jsonify({"error": "Acceso no autorizado"}), 403
    
    seccion = Seccion.query.get(seccion_id)
    if not seccion or seccion.profesor_id != user_id:
        return jsonify({"error": "No tienes acceso a esta sección"}), 403
    
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import inch
    from io import BytesIO
    
    asignatura = Asignatura.query.get(seccion.asignatura_id)
    
    # Get enrolled students with attendance
    inscripciones = Inscripcion.query.filter_by(seccion_id=seccion_id, estado='activa').all()
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=18, textColor=colors.HexColor('#39E079'), spaceAfter=30)
    elements.append(Paragraph(f"Lista de Asistencia - {asignatura.nombre if asignatura else seccion.codigo}", title_style))
    
    # Info table
    day_names = ['', 'Lunes', 'Martes', 'Miercoles', 'Jueves', 'Viernes', 'Sabado', 'Domingo']
    info_data = [
        ['Materia:', asignatura.nombre if asignatura else seccion.codigo],
        ['Carrera:', 'Ing. en Sistemas'],
        ['Profesor:', f"{user.primer_nombre or ''} {user.primer_apellido or ''}".strip()],
        ['Horario:', f"{day_names[seccion.dia]} {seccion.start_time.strftime('%H:%M')} - {seccion.end_time.strftime('%H:%M')}"],
        ['Fecha:', date.today().strftime('%d/%m/%Y')]
    ]
    info_table = Table(info_data, colWidths=[2*inch, 4*inch])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#122017')),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.white),
        ('TEXTCOLOR', (1, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey)
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Group students by status
    presentes = []
    ausentes = []
    justificados = []
    
    for insc in inscripciones:
        alumno = User.query.get(insc.estudiante_id)
        if not alumno:
            continue
        
        nombre = f"{alumno.primer_nombre or ''} {alumno.primer_apellido or ''}".strip() or alumno.username
        
        # Get latest status
        asistencia = AsistenciaManual.query.filter_by(
            seccion_id=seccion_id,
            alumno_id=alumno.id
        ).order_by(AsistenciaManual.fecha.desc()).first()
        
        estado = asistencia.estado if asistencia else 'ausente'
        
        if estado == 'presente':
            presentes.append([str(len(presentes) + 1), nombre, alumno.ci or 'S/N'])
        elif estado == 'justificado':
            justificados.append([str(len(justificados) + 1), nombre, alumno.ci or 'S/N'])
        else:
            ausentes.append([str(len(ausentes) + 1), nombre, alumno.ci or 'S/N'])
    
    # Presentes table
    if presentes:
        elements.append(Paragraph(f"<b>Alumnos Presentes ({len(presentes)})</b>", styles['Heading2']))
        presente_data = [['#', 'Nombre Completo', 'Cedula']] + presentes
        presente_table = Table(presente_data, colWidths=[0.5*inch, 3.5*inch, 2*inch])
        presente_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#39E079')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')])
        ]))
        elements.append(presente_table)
        elements.append(Spacer(1, 0.3*inch))
    
    # Justificados table
    if justificados:
        elements.append(Paragraph(f"<b>Alumnos Justificados ({len(justificados)})</b>", styles['Heading2']))
        justificado_data = [['#', 'Nombre Completo', 'Cedula']] + justificados
        justificado_table = Table(justificado_data, colWidths=[0.5*inch, 3.5*inch, 2*inch])
        justificado_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FFA500')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')])
        ]))
        elements.append(justificado_table)
        elements.append(Spacer(1, 0.3*inch))
    
    # Ausentes table
    if ausentes:
        elements.append(Paragraph(f"<b>Alumnos Ausentes ({len(ausentes)})</b>", styles['Heading2']))
        ausente_data = [['#', 'Nombre Completo', 'Cedula']] + ausentes
        ausente_table = Table(ausente_data, colWidths=[0.5*inch, 3.5*inch, 2*inch])
        ausente_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FF0000')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')])
        ]))
        elements.append(ausente_table)
    
    doc.build(elements)
    buffer.seek(0)
    
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=asistencia_{seccion.codigo}.pdf'
    return response


@profesor_api_bp.route('/reporte/<int:seccion_id>/excel', methods=['GET'])
@jwt_required()
def descargar_reporte_excel(seccion_id):
    """
    GET /api/profesor/reporte/<seccion_id>/excel
    Downloads Excel attendance report
    """
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    
    if not user or user.role != 'profesor':
        return jsonify({"error": "Acceso no autorizado"}), 403
    
    seccion = Seccion.query.get(seccion_id)
    if not seccion or seccion.profesor_id != user_id:
        return jsonify({"error": "No tienes acceso a esta sección"}), 403
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    
    asignatura = Asignatura.query.get(seccion.asignatura_id)
    inscripciones = Inscripcion.query.filter_by(seccion_id=seccion_id, estado='activa').all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Asistencia"
    
    # Headers
    headers = ['#', 'Nombre Completo', 'Cedula', 'Estado']
    ws.append(headers)
    
    # Header style
    header_fill = PatternFill(start_color='39E079', end_color='39E079', fill_type='solid')
    header_font = Font(bold=True, color='000000')
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data
    for idx, insc in enumerate(inscripciones, 1):
        alumno = User.query.get(insc.estudiante_id)
        if not alumno:
            continue
        
        nombre = f"{alumno.primer_nombre or ''} {alumno.primer_apellido or ''}".strip() or alumno.username
        
        # Get latest status
        asistencia = AsistenciaManual.query.filter_by(
            seccion_id=seccion_id,
            alumno_id=alumno.id
        ).order_by(AsistenciaManual.fecha.desc()).first()
        
        estado = asistencia.estado.capitalize() if asistencia else 'Sin marcar'
        ws.append([idx, nombre, alumno.ci or 'S/N', estado])
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=asistencia_{seccion.codigo}.xlsx'
    return response
