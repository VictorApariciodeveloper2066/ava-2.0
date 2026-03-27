from flask import Blueprint, render_template, session, redirect, url_for, request
from backend.models import User, Course, User_course, Asistencia, HistorialAsistencia, DetalleAsistencia
from backend.extensions import db
from datetime import date, timedelta, datetime
import pytz

front_bp = Blueprint('front', __name__)

@front_bp.route('/')
def index():
    if 'username' in session:
        user = User.query.filter_by(username=session['username']).first()
        if user and user.role in ('estudiante', 'profesor'):
            return redirect(url_for('front.dashboard'))
        return redirect(url_for('front.predashboard'))
    return render_template('index.html')

@front_bp.route('/dashboard')
def dashboard():
    if 'username' not in session:
        return redirect(url_for('front.index'))
    user = User.query.filter_by(username=session['username']).first()
    if not user:
        return redirect(url_for('front.index'))
    if user.role not in ('estudiante', 'profesor'):
        return redirect(url_for('front.predashboard'))
    
    # Verificar si el perfil está completo
    if not user.primer_nombre or not user.primer_apellido or not user.ci:
        return redirect(url_for('front.completar_perfil'))
    
    # Para estudiantes, verificar que tengan carrera
    if user.role == 'estudiante' and not user.carrera_id:
        return redirect(url_for('front.completar_perfil'))
    
    # Verificar si tiene materias inscritas
    if user.role == 'estudiante':
        tiene_materias = db.session.query(User_course).filter_by(user_id=user.id).first()
        if not tiene_materias:
            return redirect(url_for('front.completar_perfil'))
    # Query courses associated with this user
    courses = db.session.query(Course).join(User_course, Course.id == User_course.course_id).filter(User_course.user_id == user.id).all()

    # Weekday names (Spanish) and organize courses by day index (1=Monday ... 7=Sunday)
    weekday_names = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']
    slots_by_day = {i: [] for i in range(1, 8)}
    
    # Usar timezone de Venezuela (UTC-4)
    venezuela_tz = pytz.timezone('America/Caracas')
    now = datetime.now(venezuela_tz)
    current_time = now.time()
    
    # Obtener formato de hora del usuario (por defecto 12h)
    time_format = user.formato_hora if hasattr(user, 'formato_hora') and user.formato_hora else '12h'
    
    for c in courses:
        # Formatear horas según preferencia del usuario
        if time_format == '24h':
            start_time_str = c.start_time.strftime('%H:%M') if c.start_time else ''
            end_time_str = c.end_time.strftime('%H:%M') if c.end_time else ''
        else:  # 12h
            start_time_str = c.start_time.strftime('%I:%M %p') if c.start_time else ''
            end_time_str = c.end_time.strftime('%I:%M %p') if c.end_time else ''
        
        item = {
            'id': c.id,
            'name': c.name,
            'aula': c.aula,
            'dia': c.dia,
            'start_time': start_time_str,
            'end_time': end_time_str
        }
        # Show attendance button only when today is the course day and current time is within start/end
        can_mark = False
        try:
            if c.start_time and c.end_time:
                # Support Course.dia stored as 1..7 (Mon..Sun) or 0..6 (Mon..Sun)
                today_idx1 = now.weekday() + 1
                if isinstance(c.dia, int):
                    if 1 <= c.dia <= 7:
                        day_matches = (today_idx1 == c.dia)
                    else:
                        day_matches = (now.weekday() == c.dia)
                else:
                    day_matches = False
                can_mark = day_matches and (c.start_time <= current_time <= c.end_time)
        except Exception:
            can_mark = False
        item['can_mark'] = can_mark
        slots_by_day.setdefault(c.dia, []).append(item)

    # If user is a profesor, build estudiantes_by_course for active classes
    estudiantes_by_course = {}
    try:
        if user.role == 'profesor' or (user.role == 'estudiante' and user.es_comandante):
            current_day = now.weekday() + 1
            for c in courses:
                # check active session
                is_active = False
                try:
                    if isinstance(c.dia, int):
                        if 1 <= c.dia <= 7:
                            day_matches = (current_day == c.dia)
                        else:
                            day_matches = (now.weekday() == c.dia)
                    else:
                        day_matches = False
                    is_active = day_matches and (c.start_time <= current_time <= c.end_time)
                except Exception:
                    is_active = False

                if is_active:
                    estudiantes = db.session.query(User).join(User_course, User.id == User_course.user_id).filter(User_course.course_id == c.id, User.role != 'profesor').all()
                    estudiantes_by_course[c.id] = estudiantes
    except Exception:
        estudiantes_by_course = {}

    # Build values expected by the template:
    # - days_names: 1-based mapping for weekdays used in the template (1..5)
    # - week_dates: mapping day index -> date string for current week
    # - today_index: current weekday as 1=Monday..7=Sunday
    # - schedule: alias for slots_by_day
    today = date.today()
    today_index = today.weekday() + 1

    # week start (Monday)
    week_start = today - timedelta(days=today.weekday())
    week_dates = {i: (week_start + timedelta(days=(i - 1))).strftime('%d %b') for i in range(1, 6)}

    # Attendance counts for the current week (Mon-Sun)
    week_end = week_start + timedelta(days=6)
    attended_count = db.session.query(Asistencia).filter(Asistencia.user_id == user.id, Asistencia.state == 'Presente', Asistencia.date >= week_start, Asistencia.date <= week_end).count()

    # Total classes this week: count user's courses that fall on weekdays (Mon-Fri)
    total_classes = 0
    for c in courses:
        try:
            if isinstance(c.dia, int):
                if 1 <= c.dia <= 5:
                    total_classes += 1
                elif 0 <= c.dia <= 4:
                    total_classes += 1
        except Exception:
            continue

    days_names = {i + 1: weekday_names[i] for i in range(5)}

    # Provide course count for the template (avoid calling InstrumentedList.count())
    course_count = len(courses)

    return render_template('dashboard.html', user=user, username=session['username'], schedule=slots_by_day, days_names=days_names, week_dates=week_dates, today_index=today_index, today=today, course_count=course_count, attended_count=attended_count, total_classes=total_classes, courses=courses, estudiantes_by_course=estudiantes_by_course)


@front_bp.route('/asistencia/<int:course_id>')
def ver_asistencia(course_id):
    if 'username' not in session:
        return redirect(url_for('front.login_page'))
    profesor = User.query.filter_by(username=session['username']).first()
    if not profesor or (profesor.role != 'profesor' and not (profesor.role == 'estudiante' and profesor.es_comandante)):
        return redirect(url_for('front.index'))

    course = Course.query.get_or_404(course_id)

    # verify class is active (optional: restrict access to class time)
    venezuela_tz = pytz.timezone('America/Caracas')
    now = datetime.now(venezuela_tz)
    try:
        current_day = now.weekday() + 1
        is_active = (isinstance(course.dia, int) and ((1 <= course.dia <= 7 and course.dia == current_day) or (0 <= course.dia <= 6 and course.dia == now.weekday())) and course.start_time <= now.time() <= course.end_time)
    except Exception:
        is_active = False

    # get enrolled estudiantes (role == estudiante)
    alumnos_inscritos = db.session.query(User).join(User_course, User.id == User_course.user_id).filter(User_course.course_id==course_id, User.role == 'estudiante').all()

    # Get today's attendance for this course and map by user_id
    today = date.today()
    asistencias_hoy = Asistencia.query.filter_by(course_id=course_id, date=today).all()
    mapa_asistencia = {a.user_id: a.state for a in asistencias_hoy}

    # Crear registros de asistencia para estudiantes que no tienen registro hoy
    for alumno in alumnos_inscritos:
        if alumno.id not in mapa_asistencia:
            nueva_asistencia = Asistencia(user_id=alumno.id, course_id=course_id, date=today, time=now.time(), state='Ausente')
            db.session.add(nueva_asistencia)
            mapa_asistencia[alumno.id] = 'Ausente'
    db.session.commit()

    # Get pending justificativos for this course
    from backend.models import Justificativo, LogAsistencia
    justificativos_pendientes = Justificativo.query.filter_by(course_id=course_id, estado='Pendiente').all()
    
    # Get logs for today
    logs_raw = LogAsistencia.query.filter_by(course_id=course_id, fecha=today).order_by(LogAsistencia.hora.desc()).all()
    
    logs_agrupados = []
    for log in logs_raw:
        modificador = User.query.get(log.modificado_por)
        if modificador:
            logs_agrupados.append({
                'modificador': modificador,
                'hora': log.hora,
                'accion': log.accion
            })

    # Build a simple list/dict for the template so the template doesn't need to access model attrs directly
    lista_final = []
    for alumno in alumnos_inscritos:
        primer_nombre = getattr(alumno, 'primer_nombre', None) or None
        primer_apellido = getattr(alumno, 'primer_apellido', None) or None
        ci = getattr(alumno, 'ci', None) or getattr(alumno, 'cedula', None) or None
        nombre_completo = f"{primer_apellido or ''}, {primer_nombre or alumno.username or ''}".strip().strip(',')
        inicial = (primer_nombre[0] if primer_nombre else (alumno.username[0] if getattr(alumno, 'username', None) else 'U')).upper()
        lista_final.append({
            'id': alumno.id,
            'nombre_completo': nombre_completo,
            'cedula': ci or 'S/N',
            'estado_actual': mapa_asistencia.get(alumno.id, 'Ausente'),
            'inicial': inicial,
            # expose some raw fields in case template needs them
            'primer_nombre': primer_nombre,
            'primer_apellido': primer_apellido,
            'username': getattr(alumno, 'username', None)
        })

    return render_template('Asistencia.html', alumnos=lista_final, course=course, profesor=profesor, now=now, is_active=is_active, justificativos_pendientes=justificativos_pendientes, logs=logs_agrupados)


@front_bp.route('/justificativos/<int:course_id>')
def ver_justificativos(course_id):
    if 'username' not in session:
        return redirect(url_for('front.login_page'))
    profesor = User.query.filter_by(username=session['username']).first()
    if not profesor or (profesor.role != 'profesor' and not (profesor.role == 'estudiante' and profesor.es_comandante)):
        return redirect(url_for('front.index'))

    course = Course.query.get_or_404(course_id)
    from backend.models import Justificativo
    justificativos = Justificativo.query.filter_by(course_id=course_id, estado='Pendiente').all()
    
    return render_template('Reportes.html', course=course, justificativos=justificativos)

@front_bp.route('/cargar_justificativo')
def cargar_justificativo():
    if 'username' not in session:
        return redirect(url_for('front.login_page'))
    user = User.query.filter_by(username=session['username']).first()
    if not user or user.role != 'estudiante':
        return redirect(url_for('front.index'))
    
    courses = db.session.query(Course).join(User_course, Course.id == User_course.course_id).filter(User_course.user_id == user.id).all()
    return render_template('carga_justificativos.html', user=user, courses=courses)

@front_bp.route('/ver_archivo/<filename>')
def ver_archivo(filename):
    if 'username' not in session:
        return redirect(url_for('front.login_page'))
    return render_template('ver_archivo.html', filename=filename)

@front_bp.route('/historial/<int:course_id>')
def ver_historial(course_id):
    if 'username' not in session:
        return redirect(url_for('front.login_page'))
    profesor = User.query.filter_by(username=session['username']).first()
    if not profesor or (profesor.role != 'profesor' and not (profesor.role == 'estudiante' and profesor.es_comandante)):
        return redirect(url_for('front.index'))

    course = Course.query.get_or_404(course_id)
    historiales = HistorialAsistencia.query.filter_by(course_id=course_id).order_by(HistorialAsistencia.fecha.desc(), HistorialAsistencia.hora.desc()).all()
    
    return render_template('Historial.html', course=course, historiales=historiales, profesor=profesor)

@front_bp.route('/historial')
def historial_general():
    if 'username' not in session:
        return redirect(url_for('front.login_page'))
    profesor = User.query.filter_by(username=session['username']).first()
    if not profesor or (profesor.role != 'profesor' and not (profesor.role == 'estudiante' and profesor.es_comandante)):
        return redirect(url_for('front.index'))

    courses = db.session.query(Course).join(User_course, Course.id == User_course.course_id).filter(User_course.user_id == profesor.id).all()
    historiales = HistorialAsistencia.query.join(Course).filter(Course.id.in_([c.id for c in courses])).order_by(HistorialAsistencia.fecha.desc(), HistorialAsistencia.hora.desc()).all()
    
    return render_template('historial_general.html', historiales=historiales, profesor=profesor, courses=courses)

@front_bp.route('/historial_alumno')
def historial_alumno():
    if 'username' not in session:
        return redirect(url_for('front.login_page'))
    user = User.query.filter_by(username=session['username']).first()
    
    if user.role == 'profesor' or (user.role == 'estudiante' and user.es_comandante):
        return redirect(url_for('front.historial_general'))
    
    # Obtener cursos del alumno
    courses = db.session.query(Course).join(User_course, Course.id == User_course.course_id).filter(User_course.user_id == user.id).all()
    course_ids = [c.id for c in courses]
    
    # Obtener historiales de los cursos del alumno
    historiales = HistorialAsistencia.query.filter(HistorialAsistencia.course_id.in_(course_ids)).order_by(HistorialAsistencia.fecha.desc(), HistorialAsistencia.hora.desc()).all()
    
    return render_template('historial_alumno.html', user=user, courses=courses, historiales=historiales)

@front_bp.route('/historial_detalle/<int:historial_id>')
def ver_historial_detalle(historial_id):
    if 'username' not in session:
        return redirect(url_for('front.login_page'))
    user = User.query.filter_by(username=session['username']).first()
    
    historial = HistorialAsistencia.query.get_or_404(historial_id)
    
    # Verificar que el usuario tenga acceso (profesor del curso, comandante o alumno inscrito)
    if user.role == 'profesor' or (user.role == 'estudiante' and user.es_comandante):
        # Verificar que esté asociado al curso
        inscrito = db.session.query(User_course).filter_by(user_id=user.id, course_id=historial.course_id).first()
        if not inscrito:
            return redirect(url_for('front.index'))
    else:
        # Verificar que el alumno esté inscrito en el curso
        inscrito = db.session.query(User_course).filter_by(user_id=user.id, course_id=historial.course_id).first()
        if not inscrito:
            return redirect(url_for('front.index'))
    
    detalles = DetalleAsistencia.query.filter_by(historial_id=historial_id).all()
    
    # Si no hay detalles, reconstruir desde Asistencia
    if not detalles:
        asistencias = Asistencia.query.filter_by(course_id=historial.course_id, date=historial.fecha).all()
        alumnos_inscritos = db.session.query(User).join(User_course, User.id == User_course.user_id).filter(User_course.course_id == historial.course_id, User.role == 'estudiante').all()
        mapa_estados = {a.user_id: a.state for a in asistencias}
        
        for alumno in alumnos_inscritos:
            estado = mapa_estados.get(alumno.id, 'Ausente')
            detalle = DetalleAsistencia(
                historial_id=historial_id,
                user_id=alumno.id,
                estado=estado
            )
            db.session.add(detalle)
        db.session.commit()
        detalles = DetalleAsistencia.query.filter_by(historial_id=historial_id).all()
    
    # Agrupar por estado
    presentes = []
    justificados = []
    ausentes = []
    
    for detalle in detalles:
        alumno = User.query.get(detalle.user_id)
        if alumno:
            alumno_data = {
                'nombre': f"{alumno.primer_nombre or ''} {alumno.primer_apellido or ''}".strip() or alumno.username,
                'ci': alumno.ci or 'S/N',
                'inicial': (alumno.primer_nombre[0] if alumno.primer_nombre else alumno.username[0]).upper()
            }
            if detalle.estado == 'Presente':
                presentes.append(alumno_data)
            elif detalle.estado == 'Justificado':
                justificados.append(alumno_data)
            else:
                ausentes.append(alumno_data)
    
    # Obtener el profesor del curso (solo profesors, no comandantes)
    profesor = db.session.query(User).join(User_course, User.id == User_course.user_id).filter(User_course.course_id == historial.course_id, User.role == 'profesor').first()
    
    # Si no hay profesor asignado, buscar cualquier profesor en el sistema como fallback
    if not profesor:
        profesor = User.query.filter_by(role='profesor').first()
    
    # Obtener logs del día
    from backend.models import LogAsistencia
    logs_raw = LogAsistencia.query.filter_by(course_id=historial.course_id, fecha=historial.fecha).order_by(LogAsistencia.hora.desc()).all()
    
    logs_agrupados = []
    for log in logs_raw:
        modificador = User.query.get(log.modificado_por)
        if modificador:
            logs_agrupados.append({
                'modificador': modificador,
                'hora': log.hora,
                'accion': log.accion
            })
    
    return render_template('historial_detalle.html', historial=historial, presentes=presentes, justificados=justificados, ausentes=ausentes, profesor=profesor, user=user, logs=logs_agrupados)

@front_bp.route('/descargar_pdf/<int:historial_id>')
def descargar_pdf(historial_id):
    from flask import make_response
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import inch
    from io import BytesIO
    
    if 'username' not in session:
        return redirect(url_for('front.login_page'))
    user = User.query.filter_by(username=session['username']).first()
    
    historial = HistorialAsistencia.query.get_or_404(historial_id)
    
    # Verificar acceso
    inscrito = db.session.query(User_course).filter_by(user_id=user.id, course_id=historial.course_id).first()
    if not inscrito:
        return redirect(url_for('front.index'))
    
    # Obtener profesor del curso
    profesor = db.session.query(User).join(User_course, User.id == User_course.user_id).filter(User_course.course_id == historial.course_id, User.role == 'profesor').first()
    if not profesor:
        profesor = user
    
    detalles = DetalleAsistencia.query.filter_by(historial_id=historial_id).all()
    
    # Si no hay detalles, reconstruir desde Asistencia
    if not detalles:
        asistencias = Asistencia.query.filter_by(course_id=historial.course_id, date=historial.fecha).all()
        alumnos_inscritos = db.session.query(User).join(User_course, User.id == User_course.user_id).filter(User_course.course_id == historial.course_id, User.role == 'estudiante').all()
        mapa_estados = {a.user_id: a.state for a in asistencias}
        
        for alumno in alumnos_inscritos:
            estado = mapa_estados.get(alumno.id, 'Ausente')
            detalle = DetalleAsistencia(
                historial_id=historial_id,
                user_id=alumno.id,
                estado=estado
            )
            db.session.add(detalle)
        db.session.commit()
        detalles = DetalleAsistencia.query.filter_by(historial_id=historial_id).all()
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=18, textColor=colors.HexColor('#39E079'), spaceAfter=30)
    elements.append(Paragraph(f"Lista de Asistencia - {historial.curso.name}", title_style))
    
    # Información del curso
    info_data = [
        ['Materia:', historial.curso.name],
        ['Carrera:', 'Ing. en Sistemas'],
        ['Profesor:', f"{profesor.primer_nombre or ''} {profesor.primer_apellido or ''}".strip() or profesor.username],
        ['Horario:', f"{historial.curso.start_time.strftime('%H:%M')} - {historial.curso.end_time.strftime('%H:%M')}"],
        ['Fecha:', historial.fecha.strftime('%d/%m/%Y')],
        ['Código de Sesión:', historial.codigo_sesion]
    ]
    info_table = Table(info_data, colWidths=[2*inch, 4*inch])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#122017')),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.white),
        ('TEXTCOLOR', (1, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey)
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Alumnos Presentes
    presentes = [d for d in detalles if d.estado == 'Presente']
    if presentes:
        elements.append(Paragraph(f"<b>Alumnos Presentes ({len(presentes)})</b>", styles['Heading2']))
        presente_data = [['#', 'Nombre Completo', 'Cédula']]
        for idx, detalle in enumerate(presentes, 1):
            alumno = User.query.get(detalle.user_id)
            if alumno:
                nombre = f"{alumno.primer_nombre or ''} {alumno.primer_apellido or ''}".strip() or alumno.username
                presente_data.append([str(idx), nombre, alumno.ci or 'S/N'])
        
        presente_table = Table(presente_data, colWidths=[0.5*inch, 3.5*inch, 2*inch])
        presente_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#39E079')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')])
        ]))
        elements.append(presente_table)
        elements.append(Spacer(1, 0.3*inch))
    
    # Alumnos Justificados
    justificados = [d for d in detalles if d.estado == 'Justificado']
    if justificados:
        elements.append(Paragraph(f"<b>Alumnos Justificados ({len(justificados)})</b>", styles['Heading2']))
        justificado_data = [['#', 'Nombre Completo', 'Cédula']]
        for idx, detalle in enumerate(justificados, 1):
            alumno = User.query.get(detalle.user_id)
            if alumno:
                nombre = f"{alumno.primer_nombre or ''} {alumno.primer_apellido or ''}".strip() or alumno.username
                justificado_data.append([str(idx), nombre, alumno.ci or 'S/N'])
        
        justificado_table = Table(justificado_data, colWidths=[0.5*inch, 3.5*inch, 2*inch])
        justificado_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FFA500')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')])
        ]))
        elements.append(justificado_table)
        elements.append(Spacer(1, 0.3*inch))
    
    # Alumnos Ausentes
    ausentes = [d for d in detalles if d.estado == 'Ausente']
    if ausentes:
        elements.append(Paragraph(f"<b>Alumnos Ausentes ({len(ausentes)})</b>", styles['Heading2']))
        ausente_data = [['#', 'Nombre Completo', 'Cédula']]
        for idx, detalle in enumerate(ausentes, 1):
            alumno = User.query.get(detalle.user_id)
            if alumno:
                nombre = f"{alumno.primer_nombre or ''} {alumno.primer_apellido or ''}".strip() or alumno.username
                ausente_data.append([str(idx), nombre, alumno.ci or 'S/N'])
        
        ausente_table = Table(ausente_data, colWidths=[0.5*inch, 3.5*inch, 2*inch])
        ausente_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FF0000')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')])
        ]))
        elements.append(ausente_table)
    
    doc.build(elements)
    buffer.seek(0)
    
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=asistencia_{historial.codigo_sesion}.pdf'
    return response

@front_bp.route('/descargar_excel/<int:historial_id>')
def descargar_excel(historial_id):
    from flask import make_response
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    
    if 'username' not in session:
        return redirect(url_for('front.login_page'))
    user = User.query.filter_by(username=session['username']).first()
    
    historial = HistorialAsistencia.query.get_or_404(historial_id)
    
    # Verificar acceso
    inscrito = db.session.query(User_course).filter_by(user_id=user.id, course_id=historial.course_id).first()
    if not inscrito:
        return redirect(url_for('front.index'))
    
    detalles = DetalleAsistencia.query.filter_by(historial_id=historial_id).all()
    
    # Si no hay detalles, reconstruir desde Asistencia
    if not detalles:
        asistencias = Asistencia.query.filter_by(course_id=historial.course_id, date=historial.fecha).all()
        alumnos_inscritos = db.session.query(User).join(User_course, User.id == User_course.user_id).filter(User_course.course_id == historial.course_id, User.role == 'estudiante').all()
        mapa_estados = {a.user_id: a.state for a in asistencias}
        
        for alumno in alumnos_inscritos:
            estado = mapa_estados.get(alumno.id, 'Ausente')
            detalle = DetalleAsistencia(
                historial_id=historial_id,
                user_id=alumno.id,
                estado=estado
            )
            db.session.add(detalle)
        db.session.commit()
        detalles = DetalleAsistencia.query.filter_by(historial_id=historial_id).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Asistencia"
    
    # Encabezados
    headers = ['#', 'Nombre Completo', 'Cédula', 'Estado']
    ws.append(headers)
    
    # Estilo de encabezado
    header_fill = PatternFill(start_color='39E079', end_color='39E079', fill_type='solid')
    header_font = Font(bold=True, color='000000')
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Datos
    for idx, detalle in enumerate(detalles, 1):
        alumno = User.query.get(detalle.user_id)
        if alumno:
            nombre = f"{alumno.primer_nombre or ''} {alumno.primer_apellido or ''}".strip() or alumno.username
            ws.append([idx, nombre, alumno.ci or 'S/N', detalle.estado])
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=asistencia_{historial.codigo_sesion}.xlsx'
    return response

@front_bp.route('/login')
def login_page():
    return render_template('login.html')


@front_bp.route('/recuperar')
def recuperar_page():
    return render_template('Recuperar.html')


@front_bp.route('/restablecer/<token>')
def restablecer_page(token):
    return render_template('Restablecer.html', token=token)

@front_bp.route('/register')
def register_page():
    return render_template('register.html')

@front_bp.route('/predashboard')
def predashboard():
    if 'username' not in session:
        return redirect(url_for('front.login_page'))
    user = User.query.filter_by(username=session['username']).first()
    if user and user.role in ('estudiante', 'profesor'):
        return redirect(url_for('front.dashboard'))
    return render_template('Predashboard.html')

@front_bp.route('/privacidad')
def privacidad():
    return render_template('privacidad.html')

@front_bp.route('/terminos')
def terminos():
    return render_template('terminos.html')

@front_bp.route('/cookies')
def cookies():
    return render_template('cookies.html')

@front_bp.route('/completar_perfil')
def completar_perfil():
    if 'username' not in session:
        return redirect(url_for('front.login_page'))
    user = User.query.filter_by(username=session['username']).first()
    if not user:
        return redirect(url_for('front.login_page'))
    
    # Obtener todas las materias (todas son de Ingeniería en Sistemas)
    courses = Course.query.all()
    
    return render_template('completar_perfil.html', user=user, courses=courses)

@front_bp.route('/set_role', methods=['POST'])
def set_role():
    if 'username' not in session:
        return redirect(url_for('front.login_page'))
    role = request.form.get('role')
    if role not in ('estudiante', 'profesor'):
        return redirect(url_for('front.predashboard'))
    user = User.query.filter_by(username=session['username']).first()
    if not user:
        return redirect(url_for('front.index'))
    if user.role not in ('estudiante', 'profesor'):
        user.role = role
        db.session.commit()
    return redirect(url_for('front.dashboard'))

@front_bp.route('/configuracion')
def configuracion():
    if 'username' not in session:
        return redirect(url_for('front.login_page'))
    user = User.query.filter_by(username=session['username']).first()
    if not user:
        return redirect(url_for('front.login_page'))
    
    # Obtener cursos disponibles y cursos del usuario
    all_courses = Course.query.all()
    user_courses = db.session.query(Course).join(User_course, Course.id == User_course.course_id).filter(User_course.user_id == user.id).all()
    user_course_ids = [c.id for c in user_courses]
    
    return render_template('configuration.html', user=user, all_courses=all_courses, user_course_ids=user_course_ids)


@front_bp.route('/gestionar_comandantes')
def gestionar_comandantes():
    if 'username' not in session:
        return redirect(url_for('front.login_page'))
    user = User.query.filter_by(username=session['username']).first()
    if not user or user.role != 'profesor':
        return redirect(url_for('front.dashboard'))
    
    try:
        # Obtener cursos del profesor
        courses = db.session.query(Course).join(User_course, Course.id == User_course.course_id).filter(User_course.user_id == user.id).all()
        
        # Obtener estudiantes por curso
        course_estudiantes = {}
        for course in courses:
            estudiantes = db.session.query(User).join(User_course, User.id == User_course.user_id).filter(User_course.course_id == course.id, User.role == 'estudiante').all()
            course_estudiantes[course.id] = estudiantes
        
        return render_template('gestionar_comandantes.html', user=user, courses=courses, course_estudiantes=course_estudiantes)
    except Exception as e:
        print(f"Error en gestionar_comandantes: {e}")
        import traceback
        traceback.print_exc()
        return redirect(url_for('front.dashboard'))
